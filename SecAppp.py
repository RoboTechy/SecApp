import sys
from PyQt6.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QLineEdit, QPushButton, QMessageBox
from PyQt6.QtCore import Qt
import mysql.connector
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

class SecurityGateApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Security Gate Management")
        self.setGeometry(100, 100, 400, 250)

        # Initialize UI
        self.init_ui()
        
        # Initialize database connection
        self.init_database()
        
        # Initialize Excel file
        self.excel_file = "gate_records.xlsx"
        self.init_excel()

    def init_ui(self):
        # Create central widget and layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        # Create input fields
        self.driver_name = QLineEdit()
        self.driver_name.setPlaceholderText("Enter Driver Name")
        self.license_plate = QLineEdit()
        self.license_plate.setPlaceholderText("Enter License Plate Number")
        self.time_period = QLineEdit()
        self.time_period.setPlaceholderText("Enter Time Period (e.g., 2 hours)")

        # Create submit button
        self.submit_button = QPushButton("Submit")
        self.submit_button.setEnabled(False)
        
        # Add widgets to layout
        layout.addWidget(self.driver_name)
        layout.addWidget(self.license_plate)
        layout.addWidget(self.time_period)
        layout.addWidget(self.submit_button)
        
        # Connect signals
        self.driver_name.textChanged.connect(self.check_inputs)
        self.license_plate.textChanged.connect(self.check_inputs)
        self.time_period.textChanged.connect(self.check_inputs)
        self.submit_button.clicked.connect(self.submit_data)

    def init_database(self):
        try:
            self.db = mysql.connector.connect(
                host="localhost",
                user="your_username",  # Replace with your MySQL username
                password="your_password",  # Replace with your MySQL password
                database="security_gate_db"
            )
            self.cursor = self.db.cursor()
            
            # Create table if it doesn't exist
            self.cursor.execute("""
                CREATE TABLE IF NOT EXISTS gate_records (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    driver_name VARCHAR(255),
                    license_plate VARCHAR(50),
                    time_period VARCHAR(50),
                    entry_time DATETIME
                )
            """)
            self.db.commit()
        except mysql.connector.Error as err:
            QMessageBox.critical(self, "Database Error", f"Failed to connect to database: {err}")
            sys.exit(1)

    def init_excel(self):
        # Create Excel file if it doesn't exist
        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Gate Records"
            ws.append(["ID", "Driver Name", "License Plate", "Time Period", "Entry Time"])
            wb.save(self.excel_file)

    def check_inputs(self):
        # Enable submit button only if all fields are filled
        all_filled = (self.driver_name.text().strip() and 
                     self.license_plate.text().strip() and 
                     self.time_period.text().strip())
        self.submit_button.setEnabled(all_filled)

    def submit_data(self):
        driver_name = self.driver_name.text().strip()
        license_plate = self.license_plate.text().strip()
        time_period = self.time_period.text().strip()
        entry_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        try:
            # Insert into MySQL
            query = """
                INSERT INTO gate_records (driver_name, license_plate, time_period, entry_time)
                VALUES (%s, %s, %s, %s)
            """
            values = (driver_name, license_plate, time_period, entry_time)
            self.cursor.execute(query, values)
            self.db.commit()

            # Get the inserted record's ID
            record_id = self.cursor.lastrowid

            # Append to Excel
            wb = load_workbook(self.excel_file)
            ws = wb.active
            ws.append([record_id, driver_name, license_plate, time_period, entry_time])
            wb.save(self.excel_file)

            # Show success message and clear inputs
            QMessageBox.information(self, "Success", "Record added successfully!")
            self.driver_name.clear()
            self.license_plate.clear()
            self.time_period.clear()

        except mysql.connector.Error as err:
            QMessageBox.critical(self, "Error", f"Failed to save record: {err}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to update Excel: {e}")

    def closeEvent(self, event):
        # Clean up database connection
        if hasattr(self, 'cursor'):
            self.cursor.close()
        if hasattr(self, 'db'):
            self.db.close()
        event.accept()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = SecurityGateApp()
    window.show()
    sys.exit(app.exec())